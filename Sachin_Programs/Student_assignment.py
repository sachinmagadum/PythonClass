import openpyxl


def student_data_process():
    global path, max_col
    path = "C:\Python_Class\Python_Programs\student_data.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    max_col = sheet_obj.max_column
    col_name = []
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row=1, column=i)
        col_name.append(cell_obj.value)
    student_dict = {}
    for i in range(2, m_row + 1):
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            student_dict[col_name[j - 1]] = cell_obj.value

        student_dict['percentage'] = (student_dict['A1'] + student_dict['A2'] + student_dict['A3'] + student_dict[
            'A4'] + student_dict['A5']) / 5
        print(student_dict)
        sheet_obj.cell(row=1, column=8).value = 'percentage'
        sheet_obj.cell(row=i, column=8).value = student_dict['percentage']

        wb_obj.save(path)


student_data_process()
