import openpyxl

# workbook = Workbook()
# sheet = workbook.active
#
# sheet["A2"] = "hello"
# sheet["B2"] = "world!"
#
# workbook.save(filename="hello.xlsx")


path = "C:\Python_Class\Python_Programs\student_data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
#11111
# cell_obj = sheet_obj.cell(row=1, column=1)
#
# # Print value of cell object
# # using the value attribute
# print(cell_obj.value)
#2222
# m_row = sheet_obj.max_row
# max_col = sheet_obj.max_column
#
# for i in range(1, m_row + 1):
#     for j in range(1,max_col+1):
#         cell_obj = sheet_obj.cell(row = i, column = j)
#         print(cell_obj.value)
#
# m_row = sheet_obj.max_row
# max_col = sheet_obj.max_column
#333333
m_row = sheet_obj.max_row
max_col = sheet_obj.max_column
student_dict = {}
col_name = []
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    col_name.append(cell_obj.value)
# print(col_name)
student_data = []
for i in range(2, m_row + 1):
    for j in range(1,max_col+1):
        cell_obj = sheet_obj.cell(row = i, column = j)
        # print(cell_obj.value)
        student_dict[col_name[j-1]] = cell_obj.value
    # print(student_dict)
    student_data.append(student_dict)
print(student_data)
